import re, requests, openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font
from openpyxl.styles import colors, Border, Side, PatternFill

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
wrap_alignment = Alignment(wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center")
bold_font = Font(bold=True)
header_font = Font(bold=True, size=14)
link_font = Font(color=colors.BLUE, underline="single")

excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.title = '게시글'

header = ['No', '게시글 제목', '댓글']
excel_sheet.append(["","",""])
excel_sheet.append([""] + header)
for cell in excel_sheet[2][1:]:
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    cell.fill = grey_fill

excel_sheet.column_dimensions['A'].width = 5
excel_sheet.column_dimensions['B'].width = 20
excel_sheet.column_dimensions['C'].width = 80
excel_sheet.column_dimensions['D'].width = 80

res = requests.get("https://davelee-fun.github.io/trial/board/news.html")
soup = BeautifulSoup(res.content, "html.parser")
data = soup.select("div.list_item.symph_row")
row_num = 3

for idx, item in enumerate(data):
    title = re.sub("[\n\t\s]", "", item.find("span", class_="subject_fixed").get_text())
    reply_num = re.sub("[\n\t\s]", "", item.find("span", class_="rSymph05").get_text())
    link = "https://davelee-fun.github.io/trial/board/" + item.find("a", class_="list_subject")['href']

    res_title = requests.get(link)
    soup_title = BeautifulSoup(res_title.content, 'html.parser')
    replies = soup_title.select("div.comment_content")

    data1 = [idx+1, title, '[' + reply_num + ']']
    excel_sheet.append([""] + data1)
    excel_sheet.cell(row=row_num, column=3).hyperlink = link

    for cell in excel_sheet[row_num][1:]:
        cell.alignment = wrap_alignment
        cell.font = bold_font
        cell.border = thin_border

    excel_sheet.cell(row=row_num, column=3).font = link_font
    row_num += 1

    for reply in replies:
        reply = re.sub("[\n\t\s]", "", reply.select_one("div.comment_view").get_text())
        excel_sheet.append([""] + ["", "", reply])
        for cell in excel_sheet[row_num][1:]:
            cell.alignment = wrap_alignment
            cell.border = thin_border

        row_num += 1

    if idx > 5:
        break

excel_file.save("mysite.xlsx")
excel_file.close()
